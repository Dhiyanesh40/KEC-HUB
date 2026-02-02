from __future__ import annotations

import os
import secrets
import csv
import io
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path

from fastapi import File, UploadFile
from fastapi import FastAPI, HTTPException
from fastapi import Form
from fastapi import Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, StreamingResponse
from bson import ObjectId
import anyio

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover
    PdfReader = None  # type: ignore

try:
    from pdfminer.high_level import extract_text as _pdfminer_extract_text  # type: ignore
except Exception:  # pragma: no cover
    _pdfminer_extract_text = None  # type: ignore

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore

from .auth_service import AuthService
from .database.db import connect_mongodb, disconnect_mongodb, get_db, get_student_validation_db, mongodb_ok
from .models import (
    ApiResponse,
    AlumniListResponse,
    AlumniPost,
    AlumniPostCreateRequest,
    AlumniPostListResponse,
    AuthUserResponse,
    EventCreateRequest,
    EventCreateResponse,
    EventItem,
    EventListResponse,
    EventRegistrationCreate,
    EventRegistrationsResponse,
    ChatMessagesResponse,
    ChatSendRequest,
    ChatThreadsResponse,
    LoginRequest,
    OpportunitiesResponse,
    OpportunityItem,
    ManagementInstructionCreateRequest,
    ManagementInstructionItem,
    ManagementInstructionListResponse,
    ManagementNoteItem,
    ManagementNoteListResponse,
    PlacementCreateRequest,
    PlacementItem,
    PlacementListResponse,
    PlacementExperienceCreateRequest,
    PlacementExperienceItem,
    PlacementExperienceListResponse,
    PlacementRound,
    ProfileResponse,
    ProfileUpdateRequest,
    RegisterRequest,
    ReferralDecisionRequest,
    ReferralListResponse,
    ReferralRequestCreate,
    ResumeAnalysisResponse,
    ResumeAnalysisResult,
    ResumeImprovement,
    SendOtpRequest,
    StudentPlacementStatusResponse,
    StudentRoundStatus,
    UserProfile,
    UserRole,
    VerifyOtpRequest,
)
from .database.repositories import (
    AlumniPostRepository,
    ChatMessageRepository,
    ChatThreadRepository,
    EventRegistrationRepository,
    EventRepository,
    ManagementInstructionRepository,
    ManagementNoteRepository,
    PlacementRepository,
    PlacementExperienceRepository,
    OtpRepository,
    ReferralRepository,
    StudentEmailRepository,
    UserRepository,
    VerifiedEmailRepository,
    make_thread_id,
)
from .email_sender import notify_referral_decision, notify_referral_request, notify_placement_round_selection
from .settings import settings

from .opportunity_extractor.extractor import OpportunityExtractor
from .opportunity_extractor.types import ProfileSignals
from .resume_analyzer import GroqResumeAnalyzer


# Global repository variables
_auth_service: AuthService | None = None
_user_repo: UserRepository | None = None
_alumni_posts: AlumniPostRepository | None = None
_referrals: ReferralRepository | None = None
_chat_threads: ChatThreadRepository | None = None
_chat_messages: ChatMessageRepository | None = None
_events: EventRepository | None = None
_event_regs: EventRegistrationRepository | None = None
_placements: PlacementRepository | None = None
_placement_experiences: PlacementExperienceRepository | None = None
_mgmt_instructions: ManagementInstructionRepository | None = None
_mgmt_notes: ManagementNoteRepository | None = None
_opportunity_extractor = OpportunityExtractor()
_resume_analyzer: GroqResumeAnalyzer | None = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifespan context manager for startup and shutdown events"""
    global _auth_service, _user_repo, _alumni_posts, _referrals
    global _chat_threads, _chat_messages, _events, _event_regs
    global _placements, _placement_experiences, _mgmt_instructions, _mgmt_notes
    global _resume_analyzer
    
    # Startup
    await connect_mongodb()
    if mongodb_ok():
        db = get_db()  # Main application database (kec_opportunities_hub)
        student_validation_db = get_student_validation_db()  # Student email validation (kec_hub)
        
        otp_repo = OtpRepository(db)
        verified_repo = VerifiedEmailRepository(db)
        user_repo = UserRepository(db)
        student_email_repo = StudentEmailRepository(student_validation_db)  # Use kec_hub for validation
        alumni_posts = AlumniPostRepository(db)
        referrals = ReferralRepository(db)
        chat_threads = ChatThreadRepository(db)
        chat_messages = ChatMessageRepository(db)
        events = EventRepository(db)
        event_regs = EventRegistrationRepository(db)
        placements = PlacementRepository(db)
        placement_experiences = PlacementExperienceRepository(db)
        mgmt_instructions = ManagementInstructionRepository(db)
        mgmt_notes = ManagementNoteRepository(db)

        await otp_repo.ensure_indexes()
        await verified_repo.ensure_indexes()
        await student_email_repo.ensure_indexes()
        await user_repo.ensure_indexes()
        await alumni_posts.ensure_indexes()
        await referrals.ensure_indexes()
        await chat_threads.ensure_indexes()
        await chat_messages.ensure_indexes()
        await events.ensure_indexes()
        await event_regs.ensure_indexes()
        await placements.ensure_indexes()
        await placement_experiences.ensure_indexes()
        await mgmt_instructions.ensure_indexes()
        await mgmt_notes.ensure_indexes()

        _auth_service = AuthService(otp_repo=otp_repo, verified_repo=verified_repo, user_repo=user_repo, student_email_repo=student_email_repo)
        _user_repo = user_repo
        _alumni_posts = alumni_posts
        _referrals = referrals
        _chat_threads = chat_threads
        _chat_messages = chat_messages
        _events = events
        _event_regs = event_regs
        _placements = placements
        _placement_experiences = placement_experiences
        _mgmt_instructions = mgmt_instructions
        _mgmt_notes = mgmt_notes
    else:
        # Backend can still start, but auth endpoints will return a clear error.
        _auth_service = None
        _user_repo = None
        _alumni_posts = None
        _referrals = None
        _chat_threads = None
        _chat_messages = None
        _events = None
        _event_regs = None
        _placements = None
        _mgmt_instructions = None
        _mgmt_notes = None

    _resume_analyzer = GroqResumeAnalyzer.from_settings()
    
    yield  # Application runs here
    
    # Shutdown
    await disconnect_mongodb()


app = FastAPI(title="KEC Opportunities Hub API", lifespan=lifespan)

_BACKEND_DIR = Path(__file__).resolve().parents[1]
_UPLOADS_DIR = _BACKEND_DIR / "uploads"
_RESUMES_DIR = _UPLOADS_DIR / "resumes"
_EVENT_POSTERS_DIR = _UPLOADS_DIR / "event_posters"
_MANAGEMENT_NOTES_DIR = _UPLOADS_DIR / "management_notes"

_RESUMES_DIR.mkdir(parents=True, exist_ok=True)
_EVENT_POSTERS_DIR.mkdir(parents=True, exist_ok=True)
_MANAGEMENT_NOTES_DIR.mkdir(parents=True, exist_ok=True)

# Serve uploaded files (resume) for development.
app.mount("/uploads", StaticFiles(directory=str(_UPLOADS_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origin_list(),
    allow_credentials=True,
    allow_methods=["*"] ,
    allow_headers=["*"],
)


def _extract_resume_text_pdf(data: bytes) -> str:
    if not data:
        return ""

    def _clean(text: str) -> str:
        text = (text or "").replace("\x00", " ")
        # Light normalization; keep newlines for readability.
        text = "\n".join([line.rstrip() for line in text.splitlines()])
        return text.strip()

    extracted = ""

    if PdfReader is not None:
        try:
            reader = PdfReader(io.BytesIO(data))
            parts: list[str] = []
            for page in reader.pages:
                try:
                    txt = page.extract_text() or ""
                except Exception:
                    txt = ""
                if txt:
                    parts.append(txt)
            extracted = _clean("\n\n".join(parts))
        except Exception:
            extracted = ""

    if (not extracted or len(extracted) < 30) and _pdfminer_extract_text is not None:
        try:
            extracted2 = _pdfminer_extract_text(io.BytesIO(data)) or ""
            extracted2 = _clean(extracted2)
            if len(extracted2) > len(extracted):
                extracted = extracted2
        except Exception:
            pass

    return extracted


def _safe_str_list(v) -> list[str]:
    if not isinstance(v, list):
        return []
    out: list[str] = []
    for item in v:
        if isinstance(item, str) and item.strip():
            out.append(item.strip())
    return out


def _to_resume_analysis_result(obj: dict) -> ResumeAnalysisResult:
    improvements_raw = obj.get("improvements")
    improvements: list[ResumeImprovement] = []
    if isinstance(improvements_raw, list):
        for it in improvements_raw:
            if not isinstance(it, dict):
                continue
            area = (it.get("area") or "").strip()
            rec = (it.get("recommendation") or "").strip()
            ex = it.get("example")
            ex_s = ex.strip() if isinstance(ex, str) and ex.strip() else None
            if area and rec:
                improvements.append(ResumeImprovement(area=area, recommendation=rec, example=ex_s))

    score = obj.get("overallFitScore")
    try:
        score_i = int(score)
    except Exception:
        score_i = 0
    score_i = max(0, min(100, score_i))

    suggested_summary = obj.get("suggestedSummary")
    if not isinstance(suggested_summary, str):
        suggested_summary = None
    else:
        suggested_summary = suggested_summary.strip() or None

    final_feedback = obj.get("finalFeedback")
    if not isinstance(final_feedback, str):
        final_feedback = None
    else:
        final_feedback = final_feedback.strip() or None

    return ResumeAnalysisResult(
        overallFitScore=score_i,
        strengths=_safe_str_list(obj.get("strengths")),
        gaps=_safe_str_list(obj.get("gaps")),
        improvements=improvements,
        missingKeywords=_safe_str_list(obj.get("missingKeywords")),
        suggestedSummary=suggested_summary,
        suggestedBullets=_safe_str_list(obj.get("suggestedBullets")),
        atsWarnings=_safe_str_list(obj.get("atsWarnings")),
        finalFeedback=final_feedback,
    )


@app.post("/resume/analyze", response_model=ResumeAnalysisResponse)
async def analyze_resume(
    email: str = Query(...),
    role: UserRole = Query("student"),
    jobDescription: str = Form(...),
    file: UploadFile = File(...),
) -> ResumeAnalysisResponse:
    if role != "student":
        return ResumeAnalysisResponse(success=False, message="Only students can use resume analysis.")

    if _user_repo is None:
        return ResumeAnalysisResponse(success=False, message="Database not ready.")

    user_doc = await _user_repo.find_public_by_email_and_role(email, role)
    if not user_doc:
        return ResumeAnalysisResponse(success=False, message="User not found.")

    if _resume_analyzer is None:
        return ResumeAnalysisResponse(success=False, message="Groq is not configured on the server.", groqEnabled=False)

    jd = (jobDescription or "").strip()
    if len(jd) < 20:
        return ResumeAnalysisResponse(success=False, message="Please provide a longer job description (at least 20 characters).", groqEnabled=True, model=_resume_analyzer.model)

    raw = await file.read()
    if raw is None:
        raw = b""

    max_bytes = 5 * 1024 * 1024
    if len(raw) > max_bytes:
        return ResumeAnalysisResponse(success=False, message="Resume too large (max 5MB).", groqEnabled=True, model=_resume_analyzer.model)

    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()

    resume_text = ""
    if filename.endswith(".pdf") or content_type == "application/pdf":
        if PdfReader is None and _pdfminer_extract_text is None:
            return ResumeAnalysisResponse(
                success=False,
                message="PDF text extraction is not available on the server. Install 'pypdf' or 'pdfminer.six' and restart the backend.",
                groqEnabled=True,
                model=_resume_analyzer.model,
            )
        resume_text = _extract_resume_text_pdf(raw)
    else:
        try:
            resume_text = raw.decode("utf-8", errors="ignore").strip()
        except Exception:
            resume_text = ""

    if not resume_text or len(resume_text) < 30:
        return ResumeAnalysisResponse(
            success=False,
            message="Could not extract readable text from the resume. Please upload a text-based PDF (not scanned) or a .txt resume.",
            groqEnabled=True,
            model=_resume_analyzer.model,
        )

    obj = await _resume_analyzer.analyze(resume_text=resume_text, job_description=jd)
    if not obj:
        return ResumeAnalysisResponse(success=False, message="Resume analysis failed. Try again later.", groqEnabled=True, model=_resume_analyzer.model)

    result = _to_resume_analysis_result(obj)

    return ResumeAnalysisResponse(
        success=True,
        message="Resume analysis generated.",
        groqEnabled=True,
        model=_resume_analyzer.model,
        result=result,
    )


def _iso(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat()


def _parse_dt(s: str) -> datetime:
    raw = (s or "").strip()
    if not raw:
        raise ValueError("Invalid datetime")
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    return datetime.fromisoformat(raw)


def _doc_id(d: dict) -> str:
    _id = d.get("_id")
    return str(_id) if _id is not None else ""


def _require_role(actual: str, expected: str) -> None:
    if (actual or "").strip().lower() != expected:
        raise ValueError(f"Role must be {expected}.")


def _to_event_item(d: dict) -> EventItem:
    start_at = d.get("startAt")
    end_at = d.get("endAt")
    created = d.get("createdAt")
    return EventItem(
        id=_doc_id(d),
        managerEmail=d.get("managerEmail"),
        title=d.get("title", ""),
        description=d.get("description", ""),
        venue=d.get("venue"),
        startAt=_iso(start_at) if isinstance(start_at, datetime) else str(start_at or ""),
        endAt=_iso(end_at) if isinstance(end_at, datetime) else (str(end_at) if end_at else None),
        allowedDepartments=d.get("allowedDepartments") or [],
        formFields=d.get("formFields") or [],
        poster=d.get("poster"),
        createdAt=_iso(created) if isinstance(created, datetime) else str(created or ""),
    )


def _to_placement_item(d: dict) -> PlacementItem:
    created = d.get("createdAt")
    rounds_data = d.get("rounds") or []
    rounds = [
        PlacementRound(
            roundNumber=r.get("roundNumber", 0),
            name=r.get("name", ""),
            description=r.get("description"),
            selectedStudents=r.get("selectedStudents", []),
            uploadedAt=r.get("uploadedAt"),
            uploadedBy=r.get("uploadedBy"),
        )
        for r in rounds_data
    ]
    return PlacementItem(
        id=_doc_id(d),
        staffEmail=d.get("staffEmail"),
        companyName=d.get("companyName", ""),
        title=d.get("title", ""),
        description=d.get("description", ""),
        instructions=d.get("instructions"),
        visitDate=d.get("visitDate"),
        applicationDeadline=d.get("applicationDeadline"),
        location=d.get("location"),
        applyUrl=d.get("applyUrl"),
        allowedDepartments=d.get("allowedDepartments") or [],
        minCgpa=d.get("minCgpa"),
        maxArrears=d.get("maxArrears"),
        resources=d.get("resources") or [],
        rounds=rounds,
        createdAt=_iso(created) if isinstance(created, datetime) else str(created or ""),
    )


def _to_instruction_item(d: dict) -> ManagementInstructionItem:
    created = d.get("createdAt")
    return ManagementInstructionItem(
        id=_doc_id(d),
        staffEmail=d.get("staffEmail"),
        title=d.get("title", ""),
        body=d.get("body", ""),
        allowedDepartments=d.get("allowedDepartments") or [],
        createdAt=_iso(created) if isinstance(created, datetime) else str(created or ""),
    )


def _to_note_item(d: dict) -> ManagementNoteItem:
    created = d.get("createdAt")
    return ManagementNoteItem(
        id=_doc_id(d),
        staffEmail=d.get("staffEmail"),
        title=d.get("title", ""),
        description=d.get("description"),
        allowedDepartments=d.get("allowedDepartments") or [],
        file=d.get("file") or {},
        createdAt=_iso(created) if isinstance(created, datetime) else str(created or ""),
    )


def _normalize_allowed_departments(raw: list[str] | None) -> tuple[list[str], list[str]]:
    allowed = [str(d).strip() for d in (raw or []) if str(d).strip()]
    if any(d.lower() in {"all", "*"} for d in allowed):
        allowed = []
    return allowed, [d.lower() for d in allowed]


def _parse_departments_csv(raw: str | None) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    if s.lower() in {"all", "*"}:
        return ["all"]
    return [p.strip() for p in s.split(",") if p.strip()]


@app.get("/health", response_model=ApiResponse)
def health() -> ApiResponse:
    return ApiResponse(success=True, message=f"ok (db: {'connected' if mongodb_ok() else 'disconnected'})")


def _is_allowed_domain(email: str) -> bool:
    domain = email.split("@")[-1].lower()
    return domain in {"kongu.edu", "kongu.ac.in"}


def _to_user_profile(user_doc: dict) -> UserProfile:
    profile = user_doc.get("profile") or {}
    # Resume url is stored as-is under profile.resume.url
    return UserProfile(
        name=user_doc.get("name", "Student"),
        email=user_doc.get("email"),
        department=user_doc.get("department", "Computer Science"),
        role=user_doc.get("role", "student"),
        roll_number=profile.get("roll_number"),
        dob=profile.get("dob"),
        personal_email=profile.get("personal_email"),
        phone_number=profile.get("phone_number"),
        cgpa=profile.get("cgpa"),
        arrears_history=profile.get("arrears_history"),
        interests=profile.get("interests") or [],
        skills=profile.get("skills") or [],
        achievements=profile.get("achievements") or [],
        blogs=profile.get("blogs") or [],
        linkedin_url=profile.get("linkedin_url"),
        github_url=profile.get("github_url"),
        leetcode_url=profile.get("leetcode_url"),
        portfolio_url=profile.get("portfolio_url"),
        projects=profile.get("projects") or [],
        resume=profile.get("resume"),
    )


def _to_opportunity_item(op) -> OpportunityItem:
    deadline = op.deadline.isoformat() if op.deadline else None
    posted = op.published_at.date().isoformat() if op.published_at else None

    return OpportunityItem(
        id=f"rt-{op.id}",
        title=op.title,
        company=op.company or "",
        type=op.kind,
        source=getattr(op, "source", ""),
        matchMethod=getattr(op, "match_method", None),
        deadline=deadline,
        description=op.excerpt or "",
        tags=op.tags or [],
        location=op.location or "",
        postedDate=posted,
        eligibility="See source page",
        requirements=[],
        sourceUrl=op.source_url,
        score=op.score,
        reasons=op.reasons or [],
    )


@app.post("/auth/send-otp", response_model=ApiResponse)
async def send_otp(payload: SendOtpRequest) -> ApiResponse:
    # Keep the same restriction as frontend to prevent abuse
    if not _is_allowed_domain(payload.email):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")

    if not mongodb_ok() or _auth_service is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        provider = await _auth_service.send_otp(payload.email)
        if provider == "smtp":
            return ApiResponse(success=True, message="Verification code sent to your email.")
        return ApiResponse(success=True, message="OTP generated (dev console mode). Check backend logs.")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))
    except Exception:
        return ApiResponse(success=False, message="OTP send failed due to server configuration. Please contact admin.")


@app.post("/auth/verify-otp", response_model=ApiResponse)
async def verify_otp(payload: VerifyOtpRequest) -> ApiResponse:
    if not mongodb_ok() or _auth_service is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        await _auth_service.verify_otp(payload.email, payload.otp)
        return ApiResponse(success=True, message="OTP verified.")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))


@app.post("/auth/register", response_model=AuthUserResponse)
async def register(payload: RegisterRequest) -> AuthUserResponse:
    if not _is_allowed_domain(payload.email):
        return AuthUserResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _auth_service is None:
        return AuthUserResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        await _auth_service.register(payload.name, payload.email, payload.password, payload.department, payload.role)
        return AuthUserResponse(
            success=True,
            message="Registration successful!",
            user=UserProfile(name=payload.name, email=payload.email, department=payload.department, role=payload.role),
        )
    except ValueError as e:
        return AuthUserResponse(success=False, message=str(e))


@app.post("/auth/login", response_model=AuthUserResponse)
async def login(payload: LoginRequest) -> AuthUserResponse:
    if not mongodb_ok() or _auth_service is None:
        raise HTTPException(status_code=503, detail="MongoDB is not connected. Start MongoDB and retry.")

    try:
        user = await _auth_service.login(payload.email, payload.password, payload.role)
        return AuthUserResponse(success=True, message="Login successful!", user=UserProfile(**user))
    except ValueError as e:
        # Authentication failures should return 401 Unauthorized, not 200 OK
        raise HTTPException(status_code=401, detail=str(e))


@app.get("/profile/{email}", response_model=ProfileResponse)
async def get_profile(email: str, role: UserRole = "student") -> ProfileResponse:
    if not _is_allowed_domain(email):
        return ProfileResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _user_repo is None:
        return ProfileResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    user_doc = await _user_repo.find_public_by_email_and_role(email, role)
    if user_doc is None:
        return ProfileResponse(success=False, message="User not found.")

    return ProfileResponse(success=True, message="ok", profile=_to_user_profile(user_doc))


@app.put("/profile/{email}", response_model=ProfileResponse)
async def update_profile(email: str, payload: ProfileUpdateRequest, role: UserRole = "student") -> ProfileResponse:
    if not _is_allowed_domain(email):
        return ProfileResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _user_repo is None:
        return ProfileResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    core_update: dict = {}
    if payload.name is not None:
        core_update["name"] = payload.name
    if payload.department is not None:
        core_update["department"] = payload.department

    profile_update = payload.model_dump(exclude_none=True)
    # Remove core fields from profile_update (stored at root)
    profile_update.pop("name", None)
    profile_update.pop("department", None)

    if core_update:
        await _user_repo.update_core_fields(email, role, core_update)

    user_doc = await _user_repo.update_profile(email, role, profile_update)
    if user_doc is None:
        return ProfileResponse(success=False, message="User not found.")

    return ProfileResponse(success=True, message="Profile updated.", profile=_to_user_profile(user_doc))


@app.post("/profile/{email}/resume", response_model=ProfileResponse)
async def upload_resume(email: str, file: UploadFile = File(...), role: UserRole = "student") -> ProfileResponse:
    if not _is_allowed_domain(email):
        return ProfileResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _user_repo is None:
        return ProfileResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    if file.filename is None or not file.filename.strip():
        return ProfileResponse(success=False, message="Invalid filename.")

    # Basic allowlist
    allowed_ext = {".pdf", ".doc", ".docx"}
    original = Path(file.filename).name
    ext = Path(original).suffix.lower()
    if ext not in allowed_ext:
        return ProfileResponse(success=False, message="Only PDF/DOC/DOCX files are allowed.")

    token = secrets.token_hex(8)
    safe_identity = f"{role}_{email}".replace("@", "_").replace(".", "_")
    stored = f"{safe_identity}_{token}{ext}"
    dest = _RESUMES_DIR / stored

    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        return ProfileResponse(success=False, message="Resume too large (max 5MB).")

    dest.write_bytes(data)

    url = f"/uploads/resumes/{stored}"
    resume_meta = {
        "originalName": original,
        "storedName": stored,
        "contentType": file.content_type or "application/octet-stream",
        "size": len(data),
        "uploadedAt": datetime.now(timezone.utc).isoformat(),
        "url": url,
    }

    user_doc = await _user_repo.update_profile(email, role, {"resume": resume_meta})
    if user_doc is None:
        return ProfileResponse(success=False, message="User not found.")

    return ProfileResponse(success=True, message="Resume uploaded.", profile=_to_user_profile(user_doc))


@app.get("/opportunities/realtime/{email}", response_model=OpportunitiesResponse)
async def realtime_opportunities(email: str, role: UserRole = "student") -> OpportunitiesResponse:
    if not _is_allowed_domain(email):
        return OpportunitiesResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _user_repo is None:
        return OpportunitiesResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    user_doc = await _user_repo.find_public_by_email_and_role(email, role)
    if user_doc is None:
        return OpportunitiesResponse(success=False, message="User not found.")

    profile = _to_user_profile(user_doc)
    signals = ProfileSignals(
        email=profile.email,
        department=profile.department,
        skills=list(profile.skills or []),
        interests=list(profile.interests or []),
    )

    try:
        ops, meta = await _opportunity_extractor.extract_with_meta(signals)
        groq_enabled = bool(getattr(_opportunity_extractor, "groq_enabled", False))
        groq_used = any("groq" in str(getattr(o, "match_method", "") or "").lower() for o in ops)

        web_meta = (meta or {}).get("web") or {}
        return OpportunitiesResponse(
            success=True,
            message="ok",
            opportunities=[_to_opportunity_item(o) for o in ops],
            generatedAt=datetime.now(timezone.utc).isoformat(),
            groqEnabled=groq_enabled,
            groqUsed=groq_used,
            webSearchEnabled=bool(web_meta.get("enabled")),
            webSearchProvider=str(web_meta.get("provider")) if web_meta.get("provider") else None,
            webSearchUsed=bool(web_meta.get("used")),
            webSearchError=str(web_meta.get("error")) if web_meta.get("error") else None,
        )
    except Exception:
        return OpportunitiesResponse(success=False, message="Failed to extract opportunities. Try again.")


@app.get("/alumni/list", response_model=AlumniListResponse)
async def list_alumni(limit: int = Query(default=50, ge=1, le=200)) -> AlumniListResponse:
    if not mongodb_ok() or _user_repo is None:
        return AlumniListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    cur = _user_repo.col.find({"role": "alumni"}, {"passwordHash": 0}).limit(int(limit))
    docs = [d async for d in cur]
    alumni = [
        {
            "name": d.get("name", "Alumni"),
            "email": d.get("email"),
            "department": d.get("department", ""),
            "role": d.get("role", "alumni"),
        }
        for d in docs
        if d.get("email")
    ]
    return AlumniListResponse(success=True, message="ok", alumni=alumni)


@app.get("/alumni/posts", response_model=AlumniPostListResponse)
async def list_alumni_posts(limit: int = Query(default=100, ge=1, le=300)) -> AlumniPostListResponse:
    if not mongodb_ok() or _alumni_posts is None:
        return AlumniPostListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    docs = await _alumni_posts.list_all(limit=limit)
    posts = [
        AlumniPost(
            id=_doc_id(d),
            alumniEmail=d.get("alumniEmail"),
            title=d.get("title", ""),
            description=d.get("description", ""),
            tags=d.get("tags") or [],
            link=d.get("link"),
            createdAt=_iso(d.get("createdAt") or datetime.now(timezone.utc)),
        )
        for d in docs
    ]
    return AlumniPostListResponse(success=True, message="ok", posts=posts)


@app.post("/placements", response_model=ApiResponse)
async def create_placement_notice(payload: PlacementCreateRequest) -> ApiResponse:
    if not _is_allowed_domain(str(payload.staffEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _placements is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        _require_role(payload.role, "management")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    staff = await _user_repo.find_public_by_email_and_role(str(payload.staffEmail), "management")
    if staff is None:
        return ApiResponse(success=False, message="Management user not found.")

    # Normalize departments: empty => all, 'all'/'*' => all
    allowed = [d.strip() for d in (payload.allowedDepartments or []) if str(d).strip()]
    if any(d.strip().lower() in {"all", "*"} for d in allowed):
        allowed = []
    allowed_lower = [d.lower() for d in allowed]

    # Convert rounds info to storage format
    rounds_data = [
        {
            "roundNumber": idx + 1,
            "name": r.name,
            "description": r.description,
            "selectedStudents": [],
            "uploadedAt": None,
            "uploadedBy": None,
        }
        for idx, r in enumerate(payload.rounds or [])
    ]

    await _placements.create(
        {
            "staffEmail": str(payload.staffEmail),
            "companyName": payload.companyName,
            "title": payload.title,
            "description": payload.description,
            "instructions": payload.instructions,
            "visitDate": payload.visitDate,
            "applicationDeadline": payload.applicationDeadline,
            "location": payload.location,
            "applyUrl": payload.applyUrl,
            "allowedDepartments": allowed,
            "allowedDepartmentsLower": allowed_lower,
            "minCgpa": float(payload.minCgpa) if payload.minCgpa is not None else None,
            "maxArrears": int(payload.maxArrears) if payload.maxArrears is not None else None,
            "resources": [r.model_dump() for r in (payload.resources or [])],
            "rounds": rounds_data,
            "createdAt": datetime.now(timezone.utc),
        }
    )

    return ApiResponse(success=True, message="Placement notice created.")


@app.get("/placements/mine/{email}", response_model=PlacementListResponse)
async def list_my_placement_notices(
    email: str,
    role: UserRole = "management",
    limit: int = Query(default=200, ge=1, le=500),
) -> PlacementListResponse:
    if not _is_allowed_domain(email):
        return PlacementListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _placements is None:
        return PlacementListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "management")
    except ValueError as e:
        return PlacementListResponse(success=False, message=str(e))

    docs = await _placements.list_by_staff(email, limit=int(limit))
    return PlacementListResponse(success=True, message="ok", notices=[_to_placement_item(d) for d in docs])


@app.get("/placements/visible/{email}", response_model=PlacementListResponse)
async def list_visible_placement_notices(
    email: str,
    role: UserRole = "student",
    limit: int = Query(default=200, ge=1, le=500),
) -> PlacementListResponse:
    if not _is_allowed_domain(email):
        return PlacementListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _placements is None or _user_repo is None:
        return PlacementListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "student")
    except ValueError as e:
        return PlacementListResponse(success=False, message=str(e))

    student = await _user_repo.find_public_by_email_and_role(email, "student")
    if student is None:
        return PlacementListResponse(success=False, message="Student not found.")

    dept = str(student.get("department") or "").strip()
    profile = student.get("profile") or {}
    cgpa = profile.get("cgpa")
    arrears = profile.get("arrears_history")

    docs = await _placements.list_visible_for_student(
        student_department=dept,
        student_cgpa=float(cgpa) if cgpa is not None else None,
        student_arrears=int(arrears) if arrears is not None else None,
        limit=int(limit),
    )
    return PlacementListResponse(success=True, message="ok", notices=[_to_placement_item(d) for d in docs])


@app.get("/placements/{notice_id}/export", response_model=None)
async def export_eligible_students_csv(
    notice_id: str,
    email: str,
    role: UserRole = "management",
) :
    if not _is_allowed_domain(email):
        return JSONResponse(
            status_code=400,
            content=ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.").model_dump(),
        )
    if not mongodb_ok() or _placements is None or _user_repo is None:
        return JSONResponse(
            status_code=503,
            content=ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.").model_dump(),
        )
    try:
        _require_role(role, "management")
    except ValueError as e:
        return JSONResponse(status_code=403, content=ApiResponse(success=False, message=str(e)).model_dump())

    notice = await _placements.get_by_id(notice_id)
    if notice is None:
        return JSONResponse(status_code=404, content=ApiResponse(success=False, message="Placement notice not found.").model_dump())

    staff_email = str(notice.get("staffEmail") or "").strip().lower()
    if staff_email != email.strip().lower():
        return JSONResponse(status_code=403, content=ApiResponse(success=False, message="Not allowed.").model_dump())

    allowed_lower = notice.get("allowedDepartmentsLower") or []
    if not isinstance(allowed_lower, list):
        allowed_lower = []
    allowed_lower = [str(x).strip().lower() for x in allowed_lower if str(x).strip()]

    min_cgpa = notice.get("minCgpa")
    max_arrears = notice.get("maxArrears")
    try:
        min_cgpa_f = float(min_cgpa) if min_cgpa is not None else None
    except Exception:
        min_cgpa_f = None
    try:
        max_arr_i = int(max_arrears) if max_arrears is not None else None
    except Exception:
        max_arr_i = None

    # Pull students (best-effort). We filter in Python to handle case-insensitive dept matching
    # without requiring schema migration.
    cur = _user_repo.col.find({"$or": [{"role": "student"}, {"role": {"$exists": False}}]}, {"passwordHash": 0})
    students = [d async for d in cur]

    rows: list[dict] = []
    for s in students:
        dept = str(s.get("department") or "").strip()
        dept_l = dept.lower()

        if allowed_lower and dept_l not in allowed_lower:
            continue

        profile = s.get("profile") or {}
        cgpa_v = profile.get("cgpa")
        arrears_v = profile.get("arrears_history")

        # If notice has a constraint and student doesn't have the value -> not eligible.
        if min_cgpa_f is not None:
            try:
                if cgpa_v is None or float(cgpa_v) < min_cgpa_f:
                    continue
            except Exception:
                continue

        if max_arr_i is not None:
            try:
                if arrears_v is None or int(arrears_v) > max_arr_i:
                    continue
            except Exception:
                continue

        resume = profile.get("resume") or {}
        rows.append(
            {
                "name": s.get("name", ""),
                "email": s.get("email", ""),
                "department": dept,
                "roll_number": profile.get("roll_number") or "",
                "cgpa": profile.get("cgpa") if profile.get("cgpa") is not None else "",
                "arrears_history": profile.get("arrears_history") if profile.get("arrears_history") is not None else "",
                "phone_number": profile.get("phone_number") or "",
                "personal_email": profile.get("personal_email") or "",
                "resume_url": resume.get("url") or "",
            }
        )

    output = io.StringIO()
    fieldnames = [
        "name",
        "email",
        "department",
        "roll_number",
        "cgpa",
        "arrears_history",
        "phone_number",
        "personal_email",
        "resume_url",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for r in rows:
        writer.writerow(r)

    company = str(notice.get("companyName") or "company")
    safe_company = "".join([c if c.isalnum() else "_" for c in company]).strip("_") or "company"
    filename = f"eligible_students_{safe_company}_{notice_id}.csv"

    csv_bytes = output.getvalue().encode("utf-8")
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@app.post("/placements/{notice_id}/round/{round_number}/upload-students", response_model=ApiResponse)
async def upload_round_students(
    notice_id: str,
    round_number: int,
    email: str = Query(...),
    role: UserRole = Query("management"),
    file: UploadFile = File(...),
) -> ApiResponse:
    """Upload CSV/Excel file with student emails or roll numbers for a specific round."""
    if not _is_allowed_domain(email):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _placements is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "management")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    notice = await _placements.get_by_id(notice_id)
    if notice is None:
        return ApiResponse(success=False, message="Placement notice not found.")

    staff_email = str(notice.get("staffEmail") or "").strip().lower()
    if staff_email != email.strip().lower():
        return ApiResponse(success=False, message="You can only upload students for your own placements.")

    # Read the uploaded file
    content = await file.read()
    if not content:
        return ApiResponse(success=False, message="Empty file uploaded.")

    filename = (file.filename or "").lower()
    student_identifiers: list[str] = []

    # Handle Excel files (.xlsx, .xls)
    if filename.endswith((".xlsx", ".xls")):
        if openpyxl is None:
            return ApiResponse(success=False, message="Excel support not available. Install openpyxl and restart the server.")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            
            # Read first row and first column to detect orientation
            headers_row = []
            for cell in sheet[1]:
                if cell.value:
                    headers_row.append(str(cell.value).strip())
            
            headers_col = []
            for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                if row and row[0]:
                    headers_col.append(str(row[0]).strip())
            
            # Determine if data is transposed
            is_transposed = False
            if len(headers_col) >= len(headers_row):
                # Check if first column contains header-like values
                col_headers_lower = [h.lower().strip() for h in headers_col[:10]]
                has_expected_headers = any(
                    h in col_headers_lower 
                    for h in ["email", "student_email", "roll_number", "rollnumber", "name", "student_name"]
                )
                if has_expected_headers:
                    is_transposed = True
            
            if is_transposed:
                # Transposed format: headers in first column, data in rows to the right
                header_to_row = {}
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                    if row and row[0]:
                        header_name = str(row[0]).strip().lower()
                        header_to_row[header_name] = row_idx
                
                # Find which row contains email, roll_number, or name
                email_row = None
                rollno_row = None
                name_row = None
                
                for h, row_idx in header_to_row.items():
                    if h in ["email", "student_email", "studentemail", "e-mail"]:
                        email_row = row_idx
                    elif h in ["roll_number", "rollnumber", "roll_no", "rollno", "roll"]:
                        rollno_row = row_idx
                    elif h in ["name", "student_name", "studentname", "full_name"]:
                        name_row = row_idx
                
                # Read data from columns (starting from column 2)
                max_col = sheet.max_column
                for col_idx in range(2, max_col + 1):
                    identifier = None
                    
                    if email_row:
                        cell_value = sheet.cell(row=email_row, column=col_idx).value
                        if cell_value:
                            identifier = str(cell_value).strip().lower()
                    elif rollno_row:
                        cell_value = sheet.cell(row=rollno_row, column=col_idx).value
                        if cell_value:
                            identifier = str(cell_value).strip()
                    elif name_row:
                        cell_value = sheet.cell(row=name_row, column=col_idx).value
                        if cell_value:
                            identifier = str(cell_value).strip()
                    
                    if identifier:
                        student_identifiers.append(identifier)
            
            else:
                # Normal format: headers in first row, data in rows below
                headers = headers_row
                
                if not headers:
                    return ApiResponse(success=False, message="Excel file has no headers in the first row.")
                
                # Find column indices
                email_col_idx = None
                rollno_col_idx = None
                name_col_idx = None
                
                for idx, h in enumerate(headers):
                    h_lower = h.lower().strip()
                    if h_lower in ["email", "student_email", "studentemail", "e-mail"]:
                        email_col_idx = idx
                    elif h_lower in ["roll_number", "rollnumber", "roll_no", "rollno", "roll"]:
                        rollno_col_idx = idx
                    elif h_lower in ["name", "student_name", "studentname", "full_name"]:
                        name_col_idx = idx
                
                # Read data rows (skip header row)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                        continue
                    
                    identifier = None
                    if email_col_idx is not None and len(row) > email_col_idx and row[email_col_idx]:
                        identifier = str(row[email_col_idx]).strip().lower()
                    elif rollno_col_idx is not None and len(row) > rollno_col_idx and row[rollno_col_idx]:
                        identifier = str(row[rollno_col_idx]).strip()
                    elif name_col_idx is not None and len(row) > name_col_idx and row[name_col_idx]:
                        identifier = str(row[name_col_idx]).strip()
                    
                    if identifier:
                        student_identifiers.append(identifier)
            
            workbook.close()
        
        except Exception as e:
            return ApiResponse(success=False, message=f"Failed to parse Excel file: {str(e)}")
    
    # Handle CSV files
    else:
        try:
            text = content.decode("utf-8-sig")  # Handle BOM
        except UnicodeDecodeError:
            try:
                text = content.decode("latin-1")
            except Exception:
                return ApiResponse(success=False, message="Could not decode file. Please ensure it's UTF-8 CSV or Excel format.")

        # Parse CSV with proper newline handling
        try:
            # Use StringIO with newline parameter to handle embedded newlines properly
            reader = csv.DictReader(io.StringIO(text, newline=''))
            headers = reader.fieldnames or []
            
            # Look for common column names (case-insensitive)
            email_col = None
            rollno_col = None
            name_col = None
            
            for h in headers:
                h_lower = h.lower().strip()
                if h_lower in ["email", "student_email", "studentemail", "e-mail"]:
                    email_col = h
                elif h_lower in ["roll_number", "rollnumber", "roll_no", "rollno", "roll"]:
                    rollno_col = h
                elif h_lower in ["name", "student_name", "studentname", "full_name"]:
                    name_col = h

            for row in reader:
                # Try email first, then roll number
                identifier = None
                if email_col and row.get(email_col):
                    val = row[email_col].strip()
                    # Remove any newlines or extra whitespace
                    val = " ".join(val.split())
                    identifier = val.lower()
                elif rollno_col and row.get(rollno_col):
                    val = row[rollno_col].strip()
                    val = " ".join(val.split())
                    identifier = val
                elif name_col and row.get(name_col):
                    val = row[name_col].strip()
                    val = " ".join(val.split())
                    identifier = val
                
                if identifier:
                    student_identifiers.append(identifier)

        except Exception as e:
            return ApiResponse(success=False, message=f"Failed to parse CSV: {str(e)}")

    if not student_identifiers:
        return ApiResponse(success=False, message="No valid student identifiers found in file. Expected columns: email, roll_number, or name.")

    # Match students in database
    student_emails: list[str] = []
    not_found: list[str] = []
    
    for identifier in student_identifiers:
        # Check if it's already an email
        if "@" in identifier:
            student = await _user_repo.find_public_by_email_and_role(identifier, "student")
            if student:
                student_emails.append(identifier)
            else:
                not_found.append(identifier)
        else:
            # Search by roll number or name
            users_col = _user_repo.col
            query = {
                "$and": [
                    {"$or": [{"role": "student"}, {"role": {"$exists": False}}]},
                    {"$or": [
                        {"profile.roll_number": identifier},
                        {"name": {"$regex": f"^{identifier}$", "$options": "i"}},
                    ]},
                ]
            }
            student = await users_col.find_one(query, {"email": 1})
            if student:
                student_emails.append(student.get("email"))
            else:
                not_found.append(identifier)

    if not student_emails:
        return ApiResponse(success=False, message=f"No students found matching the uploaded data. Unmatched: {', '.join(not_found[:5])}")

    # Update the round with selected students
    rounds = notice.get("rounds") or []
    round_found = False
    
    for r in rounds:
        if r.get("roundNumber") == round_number:
            r["selectedStudents"] = student_emails
            r["uploadedAt"] = _iso(datetime.now(timezone.utc))
            r["uploadedBy"] = email
            round_found = True
            break
    
    if not round_found:
        return ApiResponse(success=False, message=f"Round {round_number} not found in this placement.")

    # Update the placement
    await _placements.col.update_one(
        {"_id": ObjectId(notice_id)},
        {"$set": {"rounds": rounds}}
    )

    # Send notifications to selected students
    company_name = notice.get("companyName", "")
    placement_title = notice.get("title", "")
    round_info = next((r for r in rounds if r.get("roundNumber") == round_number), None)
    round_name = round_info.get("name", f"Round {round_number}") if round_info else f"Round {round_number}"

    # Send emails asynchronously (non-blocking)
    for student_email in student_emails:
        try:
            await anyio.to_thread.run_sync(
                notify_placement_round_selection,
                student_email,
                company_name,
                placement_title,
                round_number,
                round_name,
            )
        except Exception:
            pass  # Continue even if email fails

    message = f"Successfully uploaded {len(student_emails)} students for {round_name}."
    if not_found:
        message += f" Could not match {len(not_found)} identifiers."
    
    return ApiResponse(success=True, message=message)


@app.get("/placements/my-selections/{email}", response_model=StudentPlacementStatusResponse)
async def get_my_placement_selections(
    email: str,
    role: UserRole = Query("student"),
) -> StudentPlacementStatusResponse:
    """Get all rounds where this student has been selected."""
    if not _is_allowed_domain(email):
        return StudentPlacementStatusResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _placements is None:
        return StudentPlacementStatusResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "student")
    except ValueError as e:
        return StudentPlacementStatusResponse(success=False, message=str(e))

    # Find all placements where student is in any round
    placements = _placements.col.find({"rounds.selectedStudents": email.lower()})
    
    selections: list[StudentRoundStatus] = []
    async for placement in placements:
        placement_id = _doc_id(placement)
        company_name = placement.get("companyName", "")
        title = placement.get("title", "")
        
        for r in placement.get("rounds", []):
            if email.lower() in [s.lower() for s in r.get("selectedStudents", [])]:
                selections.append(
                    StudentRoundStatus(
                        placementId=placement_id,
                        companyName=company_name,
                        title=title,
                        roundNumber=r.get("roundNumber", 0),
                        roundName=r.get("name", ""),
                        notifiedAt=r.get("uploadedAt", ""),
                    )
                )
    
    return StudentPlacementStatusResponse(
        success=True,
        message=f"Found {len(selections)} round selection(s).",
        selections=selections,
    )


@app.post("/management/instructions", response_model=ApiResponse)
async def create_management_instruction(payload: ManagementInstructionCreateRequest) -> ApiResponse:
    if not _is_allowed_domain(str(payload.staffEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _mgmt_instructions is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        _require_role(payload.role, "management")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    staff = await _user_repo.find_public_by_email_and_role(str(payload.staffEmail), "management")
    if staff is None:
        return ApiResponse(success=False, message="Management user not found.")

    allowed, allowed_lower = _normalize_allowed_departments(payload.allowedDepartments)

    await _mgmt_instructions.create(
        {
            "staffEmail": str(payload.staffEmail),
            "title": payload.title,
            "body": payload.body,
            "allowedDepartments": allowed,
            "allowedDepartmentsLower": allowed_lower,
            "createdAt": datetime.now(timezone.utc),
        }
    )
    return ApiResponse(success=True, message="Instruction posted.")


@app.get("/management/instructions/mine/{email}", response_model=ManagementInstructionListResponse)
async def list_my_management_instructions(
    email: str,
    role: UserRole = "management",
    limit: int = Query(default=200, ge=1, le=500),
) -> ManagementInstructionListResponse:
    if not _is_allowed_domain(email):
        return ManagementInstructionListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _mgmt_instructions is None:
        return ManagementInstructionListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "management")
    except ValueError as e:
        return ManagementInstructionListResponse(success=False, message=str(e))

    docs = await _mgmt_instructions.list_by_staff(email, limit=int(limit))
    return ManagementInstructionListResponse(success=True, message="ok", items=[_to_instruction_item(d) for d in docs])


@app.get("/management/instructions/visible/{email}", response_model=ManagementInstructionListResponse)
async def list_visible_management_instructions(
    email: str,
    role: UserRole = "student",
    limit: int = Query(default=200, ge=1, le=500),
) -> ManagementInstructionListResponse:
    if not _is_allowed_domain(email):
        return ManagementInstructionListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _mgmt_instructions is None or _user_repo is None:
        return ManagementInstructionListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "student")
    except ValueError as e:
        return ManagementInstructionListResponse(success=False, message=str(e))

    student = await _user_repo.find_public_by_email_and_role(email, "student")
    if student is None:
        return ManagementInstructionListResponse(success=False, message="Student not found.")

    dept = str(student.get("department") or "").strip()
    docs = await _mgmt_instructions.list_visible_for_department(dept, limit=int(limit))
    return ManagementInstructionListResponse(success=True, message="ok", items=[_to_instruction_item(d) for d in docs])


@app.post("/management/notes/upload", response_model=ApiResponse)
async def upload_management_note(
    email: str,
    role: UserRole = "management",
    title: str = Form(...),
    description: str = Form(""),
    allowedDepartments: str = Form("all"),
    file: UploadFile = File(...),
) -> ApiResponse:
    if not _is_allowed_domain(email):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _mgmt_notes is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "management")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    staff = await _user_repo.find_public_by_email_and_role(email, "management")
    if staff is None:
        return ApiResponse(success=False, message="Management user not found.")

    if file.filename is None or not file.filename.strip():
        return ApiResponse(success=False, message="Invalid filename.")

    original = Path(file.filename).name
    ext = Path(original).suffix.lower()
    allowed_ext = {".pdf", ".png", ".jpg", ".jpeg"}
    if ext not in allowed_ext:
        return ApiResponse(success=False, message="Only PDF/PNG/JPG/JPEG files are allowed.")

    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        return ApiResponse(success=False, message="File too large (max 10MB).")

    token = secrets.token_hex(8)
    stored = f"mgmt_{email.replace('@','_').replace('.','_')}_{token}{ext}".replace("/", "_").replace("\\", "_")
    dest = _MANAGEMENT_NOTES_DIR / stored
    dest.write_bytes(data)

    url = f"/uploads/management_notes/{stored}"
    file_meta = {
        "originalName": original,
        "storedName": stored,
        "contentType": file.content_type or "application/octet-stream",
        "size": len(data),
        "uploadedAt": datetime.now(timezone.utc).isoformat(),
        "url": url,
    }

    raw_depts = _parse_departments_csv(allowedDepartments)
    allowed, allowed_lower = _normalize_allowed_departments(raw_depts)

    await _mgmt_notes.create(
        {
            "staffEmail": email,
            "title": str(title).strip(),
            "description": str(description).strip() or None,
            "allowedDepartments": allowed,
            "allowedDepartmentsLower": allowed_lower,
            "file": file_meta,
            "createdAt": datetime.now(timezone.utc),
        }
    )

    return ApiResponse(success=True, message="Note uploaded.")


@app.get("/management/notes/mine/{email}", response_model=ManagementNoteListResponse)
async def list_my_management_notes(
    email: str,
    role: UserRole = "management",
    limit: int = Query(default=200, ge=1, le=500),
) -> ManagementNoteListResponse:
    if not _is_allowed_domain(email):
        return ManagementNoteListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _mgmt_notes is None:
        return ManagementNoteListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "management")
    except ValueError as e:
        return ManagementNoteListResponse(success=False, message=str(e))

    docs = await _mgmt_notes.list_by_staff(email, limit=int(limit))
    return ManagementNoteListResponse(success=True, message="ok", items=[_to_note_item(d) for d in docs])


@app.get("/management/notes/visible/{email}", response_model=ManagementNoteListResponse)
async def list_visible_management_notes(
    email: str,
    role: UserRole = "student",
    limit: int = Query(default=200, ge=1, le=500),
) -> ManagementNoteListResponse:
    if not _is_allowed_domain(email):
        return ManagementNoteListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _mgmt_notes is None or _user_repo is None:
        return ManagementNoteListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "student")
    except ValueError as e:
        return ManagementNoteListResponse(success=False, message=str(e))

    student = await _user_repo.find_public_by_email_and_role(email, "student")
    if student is None:
        return ManagementNoteListResponse(success=False, message="Student not found.")

    dept = str(student.get("department") or "").strip()
    docs = await _mgmt_notes.list_visible_for_department(dept, limit=int(limit))
    return ManagementNoteListResponse(success=True, message="ok", items=[_to_note_item(d) for d in docs])


@app.post("/events", response_model=EventCreateResponse)
async def create_event(payload: EventCreateRequest) -> EventCreateResponse:
    if not _is_allowed_domain(str(payload.managerEmail)):
        return EventCreateResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None or _user_repo is None:
        return EventCreateResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        _require_role(payload.role, "event_manager")
    except ValueError as e:
        return EventCreateResponse(success=False, message=str(e))

    mgr = await _user_repo.find_public_by_email_and_role(str(payload.managerEmail), "event_manager")
    if mgr is None:
        return EventCreateResponse(success=False, message="Event manager user not found.")

    # Check for duplicate event title
    title_exists = await _events.exists_by_title_and_manager(payload.title, str(payload.managerEmail))
    if title_exists:
        return EventCreateResponse(success=False, message=f"You already have an event with the title '{payload.title}'. Please use a different title.")

    try:
        start_dt = _parse_dt(payload.startAt)
        end_dt = _parse_dt(payload.endAt) if payload.endAt else None
        if end_dt and end_dt < start_dt:
            return EventCreateResponse(success=False, message="endAt must be after startAt.")
    except ValueError:
        return EventCreateResponse(success=False, message="Invalid startAt/endAt datetime. Use ISO format.")

    # Normalize departments: empty => all, 'all'/'*' => all
    allowed = [d.strip() for d in (payload.allowedDepartments or []) if str(d).strip()]
    if any(d.strip().lower() in {"all", "*"} for d in allowed):
        allowed = []
    allowed_lower = [d.lower() for d in allowed]

    # Validate form fields
    fields = payload.formFields or []
    seen_keys: set[str] = set()
    for f in fields:
        key = str(f.key)
        if key in seen_keys:
            return EventCreateResponse(success=False, message=f"Duplicate form field key: {key}")
        seen_keys.add(key)
        if f.type == "select" and not (f.options and len(f.options) > 0):
            return EventCreateResponse(success=False, message=f"Field '{key}' is select but has no options.")

    event_id = await _events.create(
        {
            "managerEmail": str(payload.managerEmail),
            "title": payload.title,
            "description": payload.description,
            "venue": payload.venue,
            "startAt": start_dt,
            "endAt": end_dt,
            "allowedDepartments": allowed,
            "allowedDepartmentsLower": allowed_lower,
            "formFields": [f.model_dump() for f in fields],
            "poster": None,
            "createdAt": datetime.now(timezone.utc),
        }
    )

    return EventCreateResponse(success=True, message="Event created.", eventId=event_id)


@app.get("/events/mine/{email}", response_model=EventListResponse)
async def list_my_events(
    email: str,
    role: UserRole = "event_manager",
    limit: int = Query(default=100, ge=1, le=300),
) -> EventListResponse:
    if not _is_allowed_domain(email):
        return EventListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None:
        return EventListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "event_manager")
    except ValueError as e:
        return EventListResponse(success=False, message=str(e))

    docs = await _events.list_by_manager(email, limit=int(limit))
    return EventListResponse(success=True, message="ok", events=[_to_event_item(d) for d in docs])


@app.get("/events/visible/{email}", response_model=EventListResponse)
async def list_visible_events(
    email: str,
    role: UserRole = "student",
    limit: int = Query(default=100, ge=1, le=300),
) -> EventListResponse:
    if not _is_allowed_domain(email):
        return EventListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None or _user_repo is None:
        return EventListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "student")
    except ValueError as e:
        return EventListResponse(success=False, message=str(e))

    student = await _user_repo.find_public_by_email_and_role(email, "student")
    if student is None:
        return EventListResponse(success=False, message="Student not found.")
    dept = str(student.get("department") or "").strip()

    docs = await _events.list_visible_for_department(dept, limit=int(limit))
    return EventListResponse(success=True, message="ok", events=[_to_event_item(d) for d in docs])


@app.put("/events/{event_id}", response_model=ApiResponse)
async def update_event(
    event_id: str,
    payload: EventCreateRequest,
) -> ApiResponse:
    if not _is_allowed_domain(str(payload.managerEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        _require_role(payload.role, "event_manager")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    # Check if event exists and belongs to this manager
    event = await _events.get_by_id(event_id)
    if event is None:
        return ApiResponse(success=False, message="Event not found.")
    if event.get("managerEmail") != str(payload.managerEmail):
        return ApiResponse(success=False, message="You can only update your own events.")

    # Check for duplicate title (excluding current event)
    title_exists = await _events.exists_by_title_and_manager(payload.title, str(payload.managerEmail), exclude_id=event_id)
    if title_exists:
        return ApiResponse(success=False, message=f"You already have another event with the title '{payload.title}'. Please use a different title.")

    try:
        start_dt = _parse_dt(payload.startAt)
        end_dt = _parse_dt(payload.endAt) if payload.endAt else None
        if end_dt and end_dt < start_dt:
            return ApiResponse(success=False, message="endAt must be after startAt.")
    except ValueError:
        return ApiResponse(success=False, message="Invalid startAt/endAt datetime. Use ISO format.")

    # Normalize departments
    allowed = [d.strip() for d in (payload.allowedDepartments or []) if str(d).strip()]
    if any(d.strip().lower() in {"all", "*"} for d in allowed):
        allowed = []
    allowed_lower = [d.lower() for d in allowed]

    # Validate form fields
    fields = payload.formFields or []
    seen_keys: set[str] = set()
    for f in fields:
        key = str(f.key)
        if key in seen_keys:
            return ApiResponse(success=False, message=f"Duplicate form field key: {key}")
        seen_keys.add(key)
        if f.type == "select" and not (f.options and len(f.options) > 0):
            return ApiResponse(success=False, message=f"Field '{key}' is select but has no options.")

    updates = {
        "title": payload.title,
        "description": payload.description,
        "venue": payload.venue,
        "startAt": start_dt,
        "endAt": end_dt,
        "allowedDepartments": allowed,
        "allowedDepartmentsLower": allowed_lower,
        "formFields": [f.model_dump() for f in fields],
        "updatedAt": datetime.now(timezone.utc),
    }

    ok = await _events.update_event(event_id, str(payload.managerEmail), updates)
    if not ok:
        return ApiResponse(success=False, message="Failed to update event.")
    return ApiResponse(success=True, message="Event updated successfully.")


@app.put("/events/{event_id}/poster", response_model=ApiResponse)
async def update_event_poster(
    event_id: str,
    email: str,
    role: UserRole = "event_manager",
    file: UploadFile = File(...),
) -> ApiResponse:
    if not _is_allowed_domain(email):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "event_manager")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    if file.filename is None or not file.filename.strip():
        return ApiResponse(success=False, message="Invalid filename.")

    original = Path(file.filename).name
    ext = Path(original).suffix.lower()
    allowed_ext = {".png", ".jpg", ".jpeg"}
    if ext not in allowed_ext:
        return ApiResponse(success=False, message="Only PNG/JPG/JPEG posters are allowed.")

    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        return ApiResponse(success=False, message="Poster too large (max 5MB).")

    # Delete old poster if exists
    event = await _events.get_by_id(event_id)
    if event and event.get("poster") and event["poster"].get("storedName"):
        old_poster = _EVENT_POSTERS_DIR / event["poster"]["storedName"]
        if old_poster.exists():
            old_poster.unlink()

    token = secrets.token_hex(8)
    stored = f"event_{event_id}_{token}{ext}".replace("/", "_").replace("\\", "_")
    dest = _EVENT_POSTERS_DIR / stored
    dest.write_bytes(data)

    poster_meta = {
        "originalName": original,
        "storedName": stored,
        "contentType": file.content_type or "image/jpeg",
        "size": len(data),
        "uploadedAt": datetime.now(timezone.utc).isoformat(),
        "url": f"/uploads/event_posters/{stored}",
    }

    ok = await _events.set_poster(event_id, email, poster_meta)
    if not ok:
        return ApiResponse(success=False, message="Event not found or not owned by this manager.")
    return ApiResponse(success=True, message="Poster updated successfully.")


@app.post("/events/{event_id}/poster", response_model=ApiResponse)
async def upload_event_poster(
    event_id: str,
    email: str,
    role: UserRole = "event_manager",
    file: UploadFile = File(...),
) -> ApiResponse:
    if not _is_allowed_domain(email):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "event_manager")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    if file.filename is None or not file.filename.strip():
        return ApiResponse(success=False, message="Invalid filename.")

    original = Path(file.filename).name
    ext = Path(original).suffix.lower()
    allowed_ext = {".png", ".jpg", ".jpeg"}
    if ext not in allowed_ext:
        return ApiResponse(success=False, message="Only PNG/JPG/JPEG posters are allowed.")

    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        return ApiResponse(success=False, message="Poster too large (max 5MB).")

    token = secrets.token_hex(8)
    stored = f"event_{event_id}_{token}{ext}".replace("/", "_").replace("\\", "_")
    dest = _EVENT_POSTERS_DIR / stored
    dest.write_bytes(data)

    poster_meta = {
        "originalName": original,
        "storedName": stored,
        "contentType": file.content_type or "image/jpeg",
        "size": len(data),
        "uploadedAt": datetime.now(timezone.utc).isoformat(),
        "url": f"/uploads/event_posters/{stored}",
    }

    ok = await _events.set_poster(event_id, email, poster_meta)
    if not ok:
        return ApiResponse(success=False, message="Event not found or not owned by this manager.")
    return ApiResponse(success=True, message="Poster uploaded.")


@app.post("/events/{event_id}/register", response_model=ApiResponse)
async def register_for_event(event_id: str, payload: EventRegistrationCreate) -> ApiResponse:
    if not _is_allowed_domain(str(payload.studentEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None or _event_regs is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(payload.studentRole, "student")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    event_doc = await _events.get_by_id(event_id)
    if event_doc is None:
        return ApiResponse(success=False, message="Event not found.")

    student = await _user_repo.find_public_by_email_and_role(str(payload.studentEmail), "student")
    if student is None:
        return ApiResponse(success=False, message="Student not found.")
    dept = str(student.get("department") or "").strip()
    dept_l = dept.lower()

    allowed_lower = event_doc.get("allowedDepartmentsLower") or []
    if isinstance(allowed_lower, list) and len(allowed_lower) > 0 and dept_l not in [str(x).lower() for x in allowed_lower]:
        return ApiResponse(success=False, message="This event is not visible for your department.")

    # Validate dynamic form answers
    answers = payload.answers or {}
    fields = event_doc.get("formFields") or []
    for f in fields:
        key = str(f.get("key") or "").strip()
        if not key:
            continue
        required = bool(f.get("required", True))
        ftype = str(f.get("type") or "text")
        label = str(f.get("label") or key)

        val = str(answers.get(key, "")).strip() if key in answers else ""
        if required and not val:
            return ApiResponse(success=False, message=f"Missing required field: {label}")

        if val and ftype == "select":
            opts = f.get("options") or []
            if isinstance(opts, list):
                cleaned_opts = [str(o).strip() for o in opts if str(o).strip()]
                if cleaned_opts and val not in cleaned_opts:
                    return ApiResponse(success=False, message=f"Invalid value for {label}.")

    try:
        event_oid = ObjectId(event_id)
    except Exception:
        return ApiResponse(success=False, message="Invalid event id.")

    if await _event_regs.exists(event_oid, str(payload.studentEmail)):
        return ApiResponse(success=False, message="You already registered for this event.")

    try:
        await _event_regs.create(
            {
                "eventId": event_oid,
                "studentEmail": str(payload.studentEmail),
                "studentRole": "student",
                "studentDepartment": dept,
                "answers": {k: str(v) for k, v in (answers or {}).items()},
                "createdAt": datetime.now(timezone.utc),
            }
        )
    except Exception:
        return ApiResponse(success=False, message="Registration failed (maybe already registered).")

    return ApiResponse(success=True, message="Registered successfully.")


@app.get("/events/{event_id}/registrations", response_model=EventRegistrationsResponse)
async def list_event_registrations(
    event_id: str,
    email: str,
    role: UserRole = "event_manager",
    limit: int = Query(default=300, ge=1, le=1000),
) -> EventRegistrationsResponse:
    if not _is_allowed_domain(email):
        return EventRegistrationsResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _events is None or _event_regs is None:
        return EventRegistrationsResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "event_manager")
    except ValueError as e:
        return EventRegistrationsResponse(success=False, message=str(e))

    event_doc = await _events.get_by_id(event_id)
    if event_doc is None:
        return EventRegistrationsResponse(success=False, message="Event not found.")
    if str(event_doc.get("managerEmail") or "").strip().lower() != email.strip().lower():
        return EventRegistrationsResponse(success=False, message="Not allowed.")

    try:
        event_oid = ObjectId(event_id)
    except Exception:
        return EventRegistrationsResponse(success=False, message="Invalid event id.")

    docs = await _event_regs.list_by_event(event_oid, limit=int(limit))
    items = []
    for d in docs:
        created = d.get("createdAt")
        items.append(
            {
                "id": _doc_id(d),
                "eventId": event_id,
                "studentEmail": d.get("studentEmail"),
                "studentDepartment": d.get("studentDepartment"),
                "answers": d.get("answers") or {},
                "createdAt": _iso(created) if isinstance(created, datetime) else "",
            }
        )
    return EventRegistrationsResponse(success=True, message="ok", registrations=items)


@app.get("/alumni/{email}/posts", response_model=AlumniPostListResponse)
async def list_posts_by_alumni(email: str, role: UserRole = "alumni", limit: int = Query(default=100, ge=1, le=300)) -> AlumniPostListResponse:
    if not _is_allowed_domain(email):
        return AlumniPostListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _alumni_posts is None:
        return AlumniPostListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "alumni")
    except ValueError as e:
        return AlumniPostListResponse(success=False, message=str(e))

    docs = await _alumni_posts.list_by_alumni(email, limit=limit)
    posts = [
        AlumniPost(
            id=_doc_id(d),
            alumniEmail=d.get("alumniEmail"),
            title=d.get("title", ""),
            description=d.get("description", ""),
            tags=d.get("tags") or [],
            link=d.get("link"),
            createdAt=_iso(d.get("createdAt") or datetime.now(timezone.utc)),
        )
        for d in docs
    ]
    return AlumniPostListResponse(success=True, message="ok", posts=posts)


@app.post("/alumni/posts", response_model=ApiResponse)
async def create_alumni_post(payload: AlumniPostCreateRequest) -> ApiResponse:
    if not _is_allowed_domain(str(payload.alumniEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _alumni_posts is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        _require_role(payload.role, "alumni")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    alumni_user = await _user_repo.find_public_by_email_and_role(str(payload.alumniEmail), "alumni")
    if alumni_user is None:
        return ApiResponse(success=False, message="Alumni user not found.")

    await _alumni_posts.create(
        {
            "alumniEmail": str(payload.alumniEmail),
            "title": payload.title,
            "description": payload.description,
            "tags": payload.tags,
            "link": str(payload.link) if payload.link else None,
            "createdAt": datetime.now(timezone.utc),
        }
    )
    return ApiResponse(success=True, message="Post created.")


@app.put("/alumni/posts/{post_id}", response_model=ApiResponse)
async def update_alumni_post(post_id: str, payload: AlumniPostCreateRequest) -> ApiResponse:
    if not _is_allowed_domain(str(payload.alumniEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _alumni_posts is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        _require_role(payload.role, "alumni")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    # Check if post exists and belongs to this alumni
    post = await _alumni_posts.get_by_id(post_id)
    if post is None:
        return ApiResponse(success=False, message="Post not found.")
    if post.get("alumniEmail") != str(payload.alumniEmail):
        return ApiResponse(success=False, message="You can only update your own posts.")

    updates = {
        "title": payload.title,
        "description": payload.description,
        "tags": payload.tags,
        "link": str(payload.link) if payload.link else None,
        "updatedAt": datetime.now(timezone.utc),
    }

    ok = await _alumni_posts.update_post(post_id, str(payload.alumniEmail), updates)
    if not ok:
        return ApiResponse(success=False, message="Failed to update post.")
    return ApiResponse(success=True, message="Post updated successfully.")


@app.post("/referrals/request", response_model=ApiResponse)
async def request_referral(payload: ReferralRequestCreate) -> ApiResponse:
    if not _is_allowed_domain(str(payload.studentEmail)) or not _is_allowed_domain(str(payload.alumniEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _referrals is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    if (payload.studentRole or "student") != "student":
        return ApiResponse(success=False, message="studentRole must be student.")

    student = await _user_repo.find_public_by_email_and_role(str(payload.studentEmail), "student")
    if student is None:
        return ApiResponse(success=False, message="Student not found.")
    alumni = await _user_repo.find_public_by_email_and_role(str(payload.alumniEmail), "alumni")
    if alumni is None:
        return ApiResponse(success=False, message="Alumni not found.")

    # Prevent duplicate referral requests for the same post.
    if payload.postId:
        try:
            exists = await _referrals.exists_for_student_alumni_post(
                str(payload.studentEmail),
                str(payload.alumniEmail),
                payload.postId,
            )
        except Exception:
            exists = False
        if exists:
            return ApiResponse(success=False, message="You already requested a referral for this post.")

    await _referrals.create(
        {
            "studentEmail": str(payload.studentEmail),
            "studentRole": "student",
            "alumniEmail": str(payload.alumniEmail),
            "alumniRole": "alumni",
            "postId": payload.postId,
            "message": payload.message,
            "status": "pending",
            "createdAt": datetime.now(timezone.utc),
            "decidedAt": None,
            "alumniNote": None,
        }
    )

    # Best-effort email notification
    post_title = None
    if payload.postId and _alumni_posts is not None:
        post = await _alumni_posts.get_by_id(payload.postId)
        if post:
            post_title = post.get("title")

    try:
        await anyio.to_thread.run_sync(
            notify_referral_request,
            str(payload.alumniEmail),
            str(payload.studentEmail),
            payload.message,
            post_title,
        )
    except Exception as e:
        # Notification failures should not block the request
        print(f"[NOTIFY] referral request email failed: {e}")

    return ApiResponse(success=True, message="Referral request sent.")


@app.get("/referrals/inbox/{email}", response_model=ReferralListResponse)
async def referral_inbox(email: str, role: UserRole = "alumni", status: str | None = None) -> ReferralListResponse:
    if not _is_allowed_domain(email):
        return ReferralListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _referrals is None:
        return ReferralListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "alumni")
    except ValueError as e:
        return ReferralListResponse(success=False, message=str(e))

    docs = await _referrals.list_for_alumni(email, status=status)
    items = []
    for d in docs:
        decided_at = d.get("decidedAt")
        items.append(
            {
                "id": _doc_id(d),
                "studentEmail": d.get("studentEmail"),
                "alumniEmail": d.get("alumniEmail"),
                "postId": d.get("postId"),
                "message": d.get("message", ""),
                "status": d.get("status", "pending"),
                "createdAt": _iso(d.get("createdAt") or datetime.now(timezone.utc)),
                "decidedAt": _iso(decided_at) if isinstance(decided_at, datetime) else None,
                "alumniNote": d.get("alumniNote"),
            }
        )
    return ReferralListResponse(success=True, message="ok", requests=items)


@app.get("/referrals/outbox/{email}", response_model=ReferralListResponse)
async def referral_outbox(email: str, role: UserRole = "student") -> ReferralListResponse:
    if not _is_allowed_domain(email):
        return ReferralListResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _referrals is None:
        return ReferralListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(role, "student")
    except ValueError as e:
        return ReferralListResponse(success=False, message=str(e))

    docs = await _referrals.list_for_student(email)
    items = []
    for d in docs:
        decided_at = d.get("decidedAt")
        items.append(
            {
                "id": _doc_id(d),
                "studentEmail": d.get("studentEmail"),
                "alumniEmail": d.get("alumniEmail"),
                "postId": d.get("postId"),
                "message": d.get("message", ""),
                "status": d.get("status", "pending"),
                "createdAt": _iso(d.get("createdAt") or datetime.now(timezone.utc)),
                "decidedAt": _iso(decided_at) if isinstance(decided_at, datetime) else None,
                "alumniNote": d.get("alumniNote"),
            }
        )
    return ReferralListResponse(success=True, message="ok", requests=items)


@app.post("/referrals/{req_id}/decide", response_model=ApiResponse)
async def decide_referral(req_id: str, payload: ReferralDecisionRequest) -> ApiResponse:
    if not _is_allowed_domain(str(payload.alumniEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _referrals is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")
    try:
        _require_role(payload.alumniRole, "alumni")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    updated = await _referrals.decide(req_id, str(payload.alumniEmail), payload.decision, datetime.now(timezone.utc), payload.note)
    if updated is None:
        return ApiResponse(success=False, message="Invalid request id.")

    # Best-effort email notification to student
    student_email = str(updated.get("studentEmail") or "")
    post_title = None
    post_id = updated.get("postId")
    if post_id and _alumni_posts is not None:
        try:
            post = await _alumni_posts.get_by_id(str(post_id))
            if post:
                post_title = post.get("title")
        except Exception:
            post_title = None

    if student_email:
        try:
            await anyio.to_thread.run_sync(
                notify_referral_decision,
                student_email,
                str(payload.alumniEmail),
                payload.decision,
                payload.note,
                post_title,
            )
        except Exception as e:
            print(f"[NOTIFY] referral decision email failed: {e}")

    return ApiResponse(success=True, message="Decision saved.")


@app.get("/chat/threads/{email}", response_model=ChatThreadsResponse)
async def chat_threads(email: str, role: UserRole = "student") -> ChatThreadsResponse:
    if not _is_allowed_domain(email):
        return ChatThreadsResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _chat_threads is None:
        return ChatThreadsResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    docs = await _chat_threads.list_for_user(email, role)
    threads = []
    for d in docs:
        parts = d.get("participants") or []
        me = f"{role}:{email}".lower()
        other = next((p for p in parts if str(p).lower() != me), None)
        if not other or ":" not in other:
            continue
        other_role, other_email = other.split(":", 1)
        upd = d.get("updatedAt")
        threads.append(
            {
                "id": d.get("_id"),
                "otherEmail": other_email,
                "otherRole": other_role,
                "lastMessage": d.get("lastMessage"),
                "lastAt": _iso(upd) if isinstance(upd, datetime) else None,
            }
        )
    return ChatThreadsResponse(success=True, message="ok", threads=threads)


@app.get("/chat/messages/{thread_id}", response_model=ChatMessagesResponse)
async def chat_messages(thread_id: str, email: str, role: UserRole) -> ChatMessagesResponse:
    if not _is_allowed_domain(email):
        return ChatMessagesResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _chat_messages is None:
        return ChatMessagesResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    docs = await _chat_messages.list_by_thread(thread_id)
    msgs = []
    for d in docs:
        created = d.get("createdAt")
        msgs.append(
            {
                "id": _doc_id(d),
                "threadId": d.get("threadId"),
                "senderEmail": d.get("senderEmail"),
                "senderRole": d.get("senderRole"),
                "text": d.get("text", ""),
                "createdAt": _iso(created) if isinstance(created, datetime) else "",
            }
        )
    return ChatMessagesResponse(success=True, message="ok", messages=msgs)


@app.post("/chat/send", response_model=ApiResponse)
async def chat_send(payload: ChatSendRequest) -> ApiResponse:
    if not _is_allowed_domain(str(payload.senderEmail)) or not _is_allowed_domain(str(payload.recipientEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _chat_threads is None or _chat_messages is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    thread_id = make_thread_id(str(payload.senderEmail), payload.senderRole, str(payload.recipientEmail), payload.recipientRole)
    now = datetime.now(timezone.utc)
    participants = sorted([
        f"{payload.senderRole}:{payload.senderEmail}".lower(),
        f"{payload.recipientRole}:{payload.recipientEmail}".lower(),
    ])

    await _chat_messages.create(
        {
            "threadId": thread_id,
            "senderEmail": str(payload.senderEmail),
            "senderRole": payload.senderRole,
            "recipientEmail": str(payload.recipientEmail),
            "recipientRole": payload.recipientRole,
            "text": payload.text,
            "createdAt": now,
        }
    )
    await _chat_threads.upsert_on_message(
        thread_id,
        participants,
        now,
        payload.text[:500],
        str(payload.senderEmail),
    )
    return ApiResponse(success=True, message=thread_id)


def _to_placement_experience_item(doc: dict) -> PlacementExperienceItem:
    return PlacementExperienceItem(
        id=str(doc["_id"]),
        studentEmail=str(doc.get("studentEmail", "")),
        studentName=doc.get("studentName"),
        studentDepartment=doc.get("studentDepartment"),
        companyName=str(doc.get("companyName", "")),
        jobRole=str(doc.get("jobRole", "")),
        interviewDate=str(doc.get("interviewDate", "")),
        rounds=doc.get("rounds", []),
        difficultyLevel=int(doc.get("difficultyLevel", 1)),
        overallExperience=str(doc.get("overallExperience", "")),
        createdAt=doc.get("createdAt", datetime.now(timezone.utc)).isoformat(),
    )


@app.post("/api/experiences", response_model=ApiResponse)
async def create_placement_experience(payload: PlacementExperienceCreateRequest) -> ApiResponse:
    if not _is_allowed_domain(str(payload.studentEmail)):
        return ApiResponse(success=False, message="Only @kongu.edu or @kongu.ac.in emails are permitted.")
    if not mongodb_ok() or _placement_experiences is None or _user_repo is None:
        return ApiResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    try:
        _require_role(payload.studentRole, "student")
    except ValueError as e:
        return ApiResponse(success=False, message=str(e))

    student = await _user_repo.find_public_by_email_and_role(str(payload.studentEmail), "student")
    if student is None:
        return ApiResponse(success=False, message="Student not found.")

    # Get student name and department for display
    student_name = student.get("name", "")
    student_dept = student.get("department", "")

    await _placement_experiences.create(
        {
            "studentEmail": str(payload.studentEmail),
            "studentName": student_name,
            "studentDepartment": student_dept,
            "companyName": payload.companyName,
            "jobRole": payload.jobRole,
            "interviewDate": payload.interviewDate,
            "rounds": [r.model_dump() for r in payload.rounds],
            "difficultyLevel": payload.difficultyLevel,
            "overallExperience": payload.overallExperience,
            "createdAt": datetime.now(timezone.utc),
        }
    )
    return ApiResponse(success=True, message="Placement experience submitted successfully!")


@app.get("/api/experiences/{company_name}", response_model=PlacementExperienceListResponse)
async def get_experiences_by_company(company_name: str, limit: int = Query(default=50, ge=1, le=200)) -> PlacementExperienceListResponse:
    if not mongodb_ok() or _placement_experiences is None:
        return PlacementExperienceListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    docs = await _placement_experiences.list_by_company(company_name, limit=int(limit))
    return PlacementExperienceListResponse(
        success=True,
        message="ok",
        experiences=[_to_placement_experience_item(d) for d in docs],
    )


@app.get("/api/experiences", response_model=PlacementExperienceListResponse)
async def list_all_experiences(limit: int = Query(default=100, ge=1, le=200)) -> PlacementExperienceListResponse:
    if not mongodb_ok() or _placement_experiences is None:
        return PlacementExperienceListResponse(success=False, message="MongoDB is not connected. Start MongoDB and retry.")

    docs = await _placement_experiences.list_all(limit=int(limit))
    return PlacementExperienceListResponse(
        success=True,
        message="ok",
        experiences=[_to_placement_experience_item(d) for d in docs],
    )

